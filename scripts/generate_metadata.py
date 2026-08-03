import ast
import csv
import json
import math
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - generation-only dependency
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
SCENARIO_SOURCE = ROOT / "scenarios"
METADATA = ROOT / "metadata"
OUTPUTS = ROOT / "outputs"
WORKBOOK_NAME = "场景元文件最终版.xlsx"

SCHEMA_VERSION = "roadtailbench.code_scene_metadata.v5"
TAXONOMY_VERSION = "roadtailbench.capability_taxonomy.v3"


EGO_ACTIONS = [
    "Overtaking",
    "Following",
    "Yielding",
    "Merging",
    "Crossing",
    "Braking",
    "Keeping",
]

HAZARD_TYPES = [
    "traffic_signs_markings",
    "separation_protection",
    "speed_control_facilities",
    "lighting_facilities",
    "road_intersection",
    "road_surface_condition",
    "road_alignment",
    "limited_sight_distance",
    "clearance_intrusion",
    "adverse_weather",
]


def compact_point_rows(points, items_per_line=4, indent=2):
    if not isinstance(points, list):
        return json.dumps(points, ensure_ascii=False, indent=indent)
    inner_indent = " " * (indent * 2)
    close_indent = " " * indent
    lines = ["["]
    for index in range(0, len(points), items_per_line):
        chunk = points[index:index + items_per_line]
        suffix = "," if index + items_per_line < len(points) else ""
        rows = ", ".join(json.dumps(point, ensure_ascii=False) for point in chunk)
        lines.append(f"{inner_indent}{rows}{suffix}")
    lines.append(f"{close_indent}]")
    return "\n".join(lines)


def json_with_compact_trajectory(data):
    if not isinstance(data, dict) or not isinstance(data.get("reference_trajectory"), list):
        return json.dumps(data, ensure_ascii=False, indent=2)
    placeholder = "__RTB_REFERENCE_TRAJECTORY__"
    compact_data = dict(data)
    compact_data["reference_trajectory"] = placeholder
    text = json.dumps(compact_data, ensure_ascii=False, indent=2)
    return text.replace(
        json.dumps(placeholder),
        compact_point_rows(data["reference_trajectory"], items_per_line=4, indent=2),
    )


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        f.write(json_with_compact_trajectory(data))
        f.write("\n")


def workbook_path():
    direct = SCENARIO_SOURCE / WORKBOOK_NAME
    if direct.exists():
        return direct
    matches = sorted(SCENARIO_SOURCE.glob("*最终版*.xlsx"))
    if matches:
        return matches[0]
    raise FileNotFoundError(f"missing final metadata workbook under {SCENARIO_SOURCE}")


def clean_cell(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def as_int(value):
    value = clean_cell(value)
    if value is None:
        return None
    return int(value)


def as_float(value):
    value = clean_cell(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        nums = parse_numbers(value)
        return max(nums) if nums else None


def as_float_list(value):
    value = clean_cell(value)
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [float(item) for item in value if item is not None]
    return parse_numbers(value)


def split_labels(value):
    value = clean_cell(value)
    if value is None:
        return []
    parts = re.split(r"[\\/、,，;；]+", str(value))
    return [part.strip() for part in parts if part and part.strip()]


def parse_numbers(value):
    if value is None:
        return []
    if isinstance(value, (int, float)):
        return [float(value)]
    return [float(match) for match in re.findall(r"[-+]?\d+(?:\.\d+)?", str(value))]


def parse_point(value):
    nums = parse_numbers(value)
    if len(nums) < 2:
        return None
    point = [round(nums[0], 3), round(nums[1], 3)]
    if len(nums) >= 3:
        point.append(round(nums[2], 3))
    return point


def parse_points(value):
    nums = parse_numbers(value)
    points = []
    for index in range(0, len(nums) - 1, 2):
        points.append([round(nums[index], 3), round(nums[index + 1], 3)])
    return points


def indexed_value(values, index, default=None):
    if not values:
        return default
    if index < len(values):
        return values[index]
    return values[-1]


def parse_trajectory(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        points = []
        for item in value:
            point = parse_point(item)
            if point:
                points.append(point)
        return points
    text = str(value).strip()
    if not text:
        return []
    try:
        parsed = ast.literal_eval(text)
    except (SyntaxError, ValueError):
        parsed = None
    if isinstance(parsed, (list, tuple)):
        points = []
        for item in parsed:
            point = parse_point(item)
            if point:
                points.append(point)
        if points:
            return points
    nums = parse_numbers(text)
    step = 3 if len(nums) % 3 == 0 else 2
    points = []
    for index in range(0, len(nums) - step + 1, step):
        point = nums[index:index + step]
        if len(point) >= 2:
            points.append([round(v, 3) for v in point])
    return points


def polyline_length(points):
    total = 0.0
    for left, right in zip(points, points[1:]):
        total += math.hypot(float(right[0]) - float(left[0]), float(right[1]) - float(left[1]))
    return round(total, 3)


def transform_from_point(point):
    data = {
        "location": {"x": round(float(point[0]), 3), "y": round(float(point[1]), 3)},
        "rotation": {"pitch": 0.0, "yaw": round(float(point[2]) if len(point) >= 3 else 0.0, 3), "roll": 0.0},
    }
    return data


def normalize_color(value):
    value = clean_cell(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        text = str(int(value))
        if len(text) == 9:
            return ",".join([text[0:3], text[3:6], text[6:9]])
        return text
    text = str(value).strip()
    return None if text.lower() in {"null", "none", "nan"} else text


def parse_weather_summary(value):
    value = clean_cell(value)
    if value is None:
        return None
    weather = {}
    for key, raw in re.findall(r"([A-Za-z_][A-Za-z0-9_]*)\s*=\s*([-+]?\d+(?:\.\d+)?)", str(value)):
        weather[key] = float(raw)
    return weather or str(value)


def load_taxonomy(wb):
    action_rows = list(wb["Ego Action"].iter_rows(min_row=2, values_only=True))
    hazard_rows = list(wb["Hazard Type"].iter_rows(min_row=2, values_only=True))
    action_items = []
    hazard_items = []
    for name, purpose, note, *_ in action_rows:
        name = clean_cell(name)
        if not name:
            continue
        action_items.append({
            "name": str(name),
            "description_zh": clean_cell(purpose),
            "note_zh": clean_cell(note),
        })
    for name, zh_name, note, *_ in hazard_rows:
        name = clean_cell(name)
        if not name:
            continue
        hazard_items.append({
            "name": str(name),
            "label_zh": clean_cell(zh_name),
            "description_zh": clean_cell(note),
        })
    return {
        "schema_version": TAXONOMY_VERSION,
        "source_workbook": str(workbook_path().relative_to(ROOT)).replace("\\", "/"),
        "ego_action": {
            "names": [item["name"] for item in action_items],
            "items": action_items,
        },
        "hazard_type": {
            "names": [item["name"] for item in hazard_items],
            "items": hazard_items,
        },
    }


def load_rows_and_taxonomy():
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to generate RoadTailBench metadata")
    wb = load_workbook(workbook_path(), read_only=True, data_only=True)
    ws = wb["场景元信息"]
    taxonomy = load_taxonomy(wb)
    hazard_zh_to_name = {
        item.get("label_zh"): item["name"]
        for item in taxonomy["hazard_type"]["items"]
        if item.get("label_zh")
    }
    rows = {}
    for excel_row, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        scenario_id = clean_cell(row[0])
        if not scenario_id or not re.fullmatch(r"RTB\d{3}", str(scenario_id)):
            continue
        trajectory = parse_trajectory(row[9])
        hazard_labels = split_labels(row[16])
        ego_actions = split_labels(row[17])
        rows[str(scenario_id)] = {
            "excel_row": excel_row,
            "scenario_id": str(scenario_id),
            "scenario_index": as_int(row[1]),
            "estimated_vehicles": as_int(row[2]),
            "estimated_pedestrians": as_int(row[3]),
            "ego_model": clean_cell(row[4]),
            "ego_color": normalize_color(row[5]),
            "ego_role_name": clean_cell(row[6]) or "ego",
            "ego_start": parse_point(row[7]),
            "ego_end": parse_point(row[8]),
            "reference_trajectory": trajectory,
            "trajectory_length_m": polyline_length(trajectory),
            "hazard_locations": parse_points(row[11]),
            "risk_radii_m": as_float_list(row[12]),
            "scene_type": clean_cell(row[13]),
            "road_speed_limit_kmh": as_float(row[14]),
            "suggested_triggered_speeds_kmh": as_float_list(row[15]),
            "hazard_labels": hazard_labels,
            "hazard_types": [hazard_zh_to_name[label] for label in hazard_labels if label in hazard_zh_to_name],
            "ego_actions": [action for action in ego_actions if action in taxonomy["ego_action"]["names"]],
            "weather_parameters": parse_weather_summary(row[18]),
        }
    return rows, taxonomy


def binary_array(names, selected):
    selected_set = set(selected)
    return {"names": names, "values": [1 if name in selected_set else 0 for name in names]}


def build_hazards(row):
    speed_limit = row["road_speed_limit_kmh"]
    primary_hazard_type = row["hazard_types"][0] if row["hazard_types"] else None
    hazards = []
    for index, center in enumerate(row["hazard_locations"]):
        radius = indexed_value(row["risk_radii_m"], index)
        triggered_speed = indexed_value(row["suggested_triggered_speeds_kmh"], index, speed_limit)
        hazard = {
            "id": f"{row['scenario_id'].lower()}_hazard_{index + 1:02d}",
            "type": primary_hazard_type,
            "types": row["hazard_types"],
            "labels": row["hazard_labels"],
            "center": center,
            "radius_m": radius,
            "reference_speed_kmh": triggered_speed if triggered_speed is not None else speed_limit,
        }
        hazards.append({key: value for key, value in hazard.items() if value not in (None, [], {})})
    return hazards


def build_metadata(row):
    role_name = str(row["ego_role_name"])
    start = row["ego_start"] or (row["reference_trajectory"][0] if row["reference_trajectory"] else None)
    end = row["ego_end"] or (row["reference_trajectory"][-1] if row["reference_trajectory"] else None)
    speed_limit = row["road_speed_limit_kmh"]
    hazards = build_hazards(row)
    reference_speed = hazards[0].get("reference_speed_kmh") if hazards else speed_limit
    metadata = {
        "schema_version": SCHEMA_VERSION,
        "scenario_id": row["scenario_id"],
        "scenario_index": row["scenario_index"],
        "town": row["scenario_id"],
        "estimated_vehicles": row["estimated_vehicles"],
        "estimated_pedestrians": row["estimated_pedestrians"],
        "ego": {
            "role_name": role_name,
            "type_id": row["ego_model"],
            "color": row["ego_color"],
            "start_match_radius_m": 8.0,
        },
        "ego_start": transform_from_point(start) if start else None,
        "ego_end": transform_from_point(end) if end else None,
        "reference_trajectory_source": "scenarios/场景元文件最终版.xlsx:Trajectory",
        "reference_trajectory_format": "x_y_yaw",
        "reference_trajectory": row["reference_trajectory"],
        "reference_trajectory_points": len(row["reference_trajectory"]),
        "trajectory_length_m": row["trajectory_length_m"],
        "trajectory_adherence_mode": "spatial",
        "scene_type": row["scene_type"],
        "speed_limit_kmh": speed_limit,
        "reference_speed_kmh": reference_speed,
        "hazard_labels": row["hazard_labels"],
        "ego_actions": row["ego_actions"],
        "weather_parameters": row["weather_parameters"],
        "capability_vector": {
            "ego_action": binary_array(EGO_ACTIONS, row["ego_actions"]),
            "hazard_type": binary_array(HAZARD_TYPES, row["hazard_types"]),
        },
        "hazards": hazards,
    }
    return metadata


def build_audit_row(metadata):
    vector = metadata["capability_vector"]
    ego = metadata.get("ego") or {}
    hazards = metadata.get("hazards") or []
    first_hazard = hazards[0] if hazards else {}
    hazard_selected = [
        name for name, value in zip(vector["hazard_type"]["names"], vector["hazard_type"]["values"]) if value
    ]
    action_selected = [
        name for name, value in zip(vector["ego_action"]["names"], vector["ego_action"]["values"]) if value
    ]
    return {
        "scenario_id": metadata["scenario_id"],
        "scenario_index": metadata.get("scenario_index"),
        "ego_model": metadata.get("ego_model") or ego.get("type_id"),
        "scene_type": metadata.get("scene_type"),
        "estimated_vehicles": metadata.get("estimated_vehicles"),
        "estimated_pedestrians": metadata.get("estimated_pedestrians"),
        "reference_trajectory_points": metadata.get("reference_trajectory_points", 0),
        "trajectory_length_m": metadata.get("trajectory_length_m", 0.0),
        "speed_limit_kmh": metadata.get("speed_limit_kmh"),
        "suggested_triggered_speed_kmh": first_hazard.get("reference_speed_kmh"),
        "risk_radius_m": first_hazard.get("radius_m"),
        "hazard_count": len(hazards),
        "hazard_types": "/".join(hazard_selected),
        "ego_actions": "/".join(action_selected),
    }


def write_audit(report):
    write_json(METADATA / "metadata_audit.json", report)
    with (METADATA / "metadata_audit.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(report[0].keys()) if report else [])
        writer.writeheader()
        writer.writerows(report)
    summary = {
        "schema_version": "roadtailbench.metadata_summary.v1",
        "scenario_count": len(report),
        "scenario_ids": [row["scenario_id"] for row in report],
        "total_trajectory_length_m": round(sum(float(row["trajectory_length_m"]) for row in report), 3),
        "scene_type_counts": {},
        "hazard_type_counts": {name: 0 for name in HAZARD_TYPES},
        "ego_action_counts": {name: 0 for name in EGO_ACTIONS},
    }
    for row in report:
        summary["scene_type_counts"][row["scene_type"]] = summary["scene_type_counts"].get(row["scene_type"], 0) + 1
        for name in split_labels(row["hazard_types"]):
            summary["hazard_type_counts"][name] += 1
        for name in split_labels(row["ego_actions"]):
            summary["ego_action_counts"][name] += 1
    write_json(METADATA / "metadata_summary.json", summary)
    write_json(OUTPUTS / "metadata_generation_report.json", report)


def main():
    rows, taxonomy = load_rows_and_taxonomy()
    if len(rows) != 100:
        raise RuntimeError(f"expected 100 scenarios in final workbook, found {len(rows)}")
    missing_scripts = [sid for sid in rows if not (SCENARIO_SOURCE / f"{sid}.py").exists()]
    if missing_scripts:
        raise RuntimeError(f"missing source scripts for: {missing_scripts}")
    METADATA.mkdir(exist_ok=True)
    OUTPUTS.mkdir(exist_ok=True)
    write_json(METADATA / "capability_taxonomy.json", taxonomy)
    report = []
    for scenario_id in sorted(rows):
        metadata = build_metadata(rows[scenario_id])
        write_json(METADATA / f"{scenario_id}.json", metadata)
        report.append(build_audit_row(metadata))
    write_audit(report)
    print(f"wrote {len(report)} metadata files from {workbook_path()}")


if __name__ == "__main__":
    main()
